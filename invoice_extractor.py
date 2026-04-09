"""
Invoice Reference Extractor
Extracts JLR reference numbers from invoice PDFs and outputs a formatted Excel spreadsheet.
"""

import re
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path

import customtkinter as ctk
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# --- Theme colours ---
BG_DARK = "#1e1e2e"
BG_CARD = "#282840"
BG_INPUT = "#313150"
ACCENT = "#7c6ff7"
ACCENT_HOVER = "#6a5ce0"
DANGER = "#e05265"
DANGER_HOVER = "#c94458"
SUCCESS = "#43d9a2"
SUCCESS_HOVER = "#36b888"
TEXT_PRIMARY = "#e0e0ef"
TEXT_SECONDARY = "#8888a8"
BORDER_COLOUR = "#3a3a5c"


def extract_references_from_pdf(pdf_path):
    """Extract the supplier name, invoice number, and all JLR reference numbers from a PDF."""
    doc = fitz.open(pdf_path)
    full_text = ""
    for page in doc:
        full_text += page.get_text()
    doc.close()

    # Extract supplier company name from "Company: XXXXX" (first occurrence, under "Supplier:")
    company_match = re.search(r"Supplier:\s*\n\s*Company:\s*(.+)", full_text)
    company_name = company_match.group(1).strip().rstrip("\xa0") if company_match else "Unknown"

    # Extract invoice number from "Invoice XXXXXXXXX"
    invoice_match = re.search(r"Invoice\s+(\d{9})", full_text)
    invoice_number = invoice_match.group(1) if invoice_match else Path(pdf_path).stem

    # Extract all JLR reference numbers (from "JLR.XXXXXXXX" patterns)
    # Order is preserved exactly as they appear in the document
    refs = re.findall(r"JLR\.(\d{8})", full_text)
    # Deduplicate while preserving document order (each ref appears twice - in line item and "Reference number:" line)
    seen = set()
    unique_refs = []
    for ref in refs:
        if ref not in seen:
            seen.add(ref)
            unique_refs.append(ref)

    return company_name, invoice_number, unique_refs


def create_spreadsheet(company_name, invoice_data, output_path):
    """Create a formatted Excel spreadsheet matching the example template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice References"

    num_invoices = len(invoice_data)
    if num_invoices == 0:
        return

    # --- Styles ---
    title_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_alignment = Alignment(horizontal="center", vertical="center")

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    data_font = Font(name="Arial", size=10)
    data_alignment = Alignment(horizontal="center")

    count_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    count_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    count_alignment = Alignment(horizontal="center")

    thin_border_side = Side(style="thin", color="BFBFBF")
    cell_border = Border(bottom=thin_border_side)
    count_border = Border(top=thin_border_side, bottom=thin_border_side)

    # --- Title Row ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_invoices)
    title_cell = ws.cell(row=1, column=1, value=f"{company_name} \u2014 Invoice Reference Numbers")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    for col in range(2, num_invoices + 1):
        c = ws.cell(row=1, column=col)
        c.fill = title_fill
    ws.row_dimensions[1].height = 30

    # --- Header Row & Data ---
    max_refs = max(len(refs) for _, refs in invoice_data)
    ws.row_dimensions[2].height = 19.5

    for col_idx, (inv_num, refs) in enumerate(invoice_data, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18 if col_idx == 1 else 13

        header_cell = ws.cell(row=2, column=col_idx, value=f"Invoice {inv_num}")
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment
        header_cell.border = cell_border

        for row_idx, ref in enumerate(refs, start=3):
            data_cell = ws.cell(row=row_idx, column=col_idx, value=ref)
            data_cell.font = data_font
            data_cell.alignment = data_alignment
            data_cell.border = cell_border

    # --- Count Row ---
    count_row = max_refs + 3
    for col_idx in range(1, num_invoices + 1):
        col_letter = get_column_letter(col_idx)
        count_cell = ws.cell(
            row=count_row,
            column=col_idx,
            value=f"=COUNTA({col_letter}3:{col_letter}{count_row - 1})",
        )
        count_cell.font = count_font
        count_cell.fill = count_fill
        count_cell.alignment = count_alignment
        count_cell.border = count_border

    wb.save(output_path)


class FileCard(ctk.CTkFrame):
    """A single file entry card with remove button."""

    def __init__(self, master, filename, on_remove, **kwargs):
        super().__init__(master, fg_color=BG_INPUT, corner_radius=8, height=40, **kwargs)
        self.pack_propagate(False)

        ctk.CTkLabel(
            self,
            text=filename,
            font=ctk.CTkFont(family="Segoe UI", size=13),
            text_color=TEXT_PRIMARY,
            anchor="w",
        ).pack(side="left", padx=(12, 0), fill="x", expand=True)

        ctk.CTkButton(
            self,
            text="\u00d7",
            width=30,
            height=28,
            font=ctk.CTkFont(size=16, weight="bold"),
            fg_color="transparent",
            hover_color=DANGER,
            text_color=TEXT_SECONDARY,
            corner_radius=6,
            command=on_remove,
        ).pack(side="right", padx=(0, 6))


class InvoiceExtractorApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Invoice Reference Extractor")
        self.geometry("680x560")
        self.minsize(500, 450)
        self.configure(fg_color=BG_DARK)

        ctk.set_appearance_mode("dark")

        self.pdf_files = []

        self._build_ui()

    def _build_ui(self):
        # ── Header bar ──
        header = ctk.CTkFrame(self, fg_color=BG_CARD, corner_radius=0, height=70)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkLabel(
            header,
            text="Invoice Reference Extractor",
            font=ctk.CTkFont(family="Segoe UI", size=22, weight="bold"),
            text_color=TEXT_PRIMARY,
        ).pack(side="left", padx=24)

        # Accent stripe under header
        ctk.CTkFrame(self, fg_color=ACCENT, height=3, corner_radius=0).pack(fill="x")

        # ── Main content ──
        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=24, pady=20)

        # ── Button row ──
        btn_frame = ctk.CTkFrame(main, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 14))

        self.add_btn = ctk.CTkButton(
            btn_frame,
            text="Add Invoices",
            command=self._add_files,
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            fg_color=ACCENT,
            hover_color=ACCENT_HOVER,
            text_color="white",
            corner_radius=8,
            height=38,
            width=150,
        )
        self.add_btn.pack(side="left")

        self.clear_btn = ctk.CTkButton(
            btn_frame,
            text="Clear All",
            command=self._clear_files,
            font=ctk.CTkFont(family="Segoe UI", size=13),
            fg_color="transparent",
            hover_color=DANGER,
            text_color=TEXT_SECONDARY,
            border_width=1,
            border_color=BORDER_COLOUR,
            corner_radius=8,
            height=38,
            width=100,
        )
        self.clear_btn.pack(side="left", padx=(10, 0))

        # File count badge
        self.count_label = ctk.CTkLabel(
            btn_frame,
            text="0 files",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=TEXT_SECONDARY,
        )
        self.count_label.pack(side="right")

        # ── File list area ──
        self.list_container = ctk.CTkScrollableFrame(
            main,
            fg_color=BG_CARD,
            corner_radius=12,
            border_width=1,
            border_color=BORDER_COLOUR,
        )
        self.list_container.pack(fill="both", expand=True, pady=(0, 14))

        # Empty state placeholder
        self.empty_label = ctk.CTkLabel(
            self.list_container,
            text="Drop your PDF invoices here\nor click \"Add Invoices\" to browse",
            font=ctk.CTkFont(family="Segoe UI", size=14),
            text_color=TEXT_SECONDARY,
            justify="center",
        )
        self.empty_label.pack(expand=True, pady=60)

        # ── Status bar ──
        self.status_label = ctk.CTkLabel(
            main,
            text="",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=TEXT_SECONDARY,
            anchor="w",
        )
        self.status_label.pack(fill="x", pady=(0, 10))

        # ── Extract button ──
        self.extract_btn = ctk.CTkButton(
            main,
            text="Extract References & Save Spreadsheet",
            command=self._extract,
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            fg_color=SUCCESS,
            hover_color=SUCCESS_HOVER,
            text_color=BG_DARK,
            corner_radius=10,
            height=48,
            state="disabled",
        )
        self.extract_btn.pack(fill="x")

        # ── Footer with GitHub credit ──
        footer = ctk.CTkFrame(main, fg_color="transparent", height=24)
        footer.pack(fill="x", pady=(10, 0))

        self.github_link = ctk.CTkLabel(
            footer,
            text="github.com/Kieronhall",
            font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color=TEXT_SECONDARY,
            cursor="hand2",
        )
        self.github_link.pack(side="right")
        self.github_link.bind("<Button-1>", lambda e: self._open_github())
        self.github_link.bind("<Enter>", lambda e: self.github_link.configure(text_color=ACCENT))
        self.github_link.bind("<Leave>", lambda e: self.github_link.configure(text_color=TEXT_SECONDARY))

    def _open_github(self):
        import webbrowser
        webbrowser.open("https://github.com/Kieronhall")

    def _refresh_file_list(self):
        """Rebuild the file card list from self.pdf_files."""
        for widget in self.list_container.winfo_children():
            widget.destroy()

        if not self.pdf_files:
            self.empty_label = ctk.CTkLabel(
                self.list_container,
                text="Drop your PDF invoices here\nor click \"Add Invoices\" to browse",
                font=ctk.CTkFont(family="Segoe UI", size=14),
                text_color=TEXT_SECONDARY,
                justify="center",
            )
            self.empty_label.pack(expand=True, pady=60)
            self.extract_btn.configure(state="disabled")
            self.count_label.configure(text="0 files")
            self.status_label.configure(text="")
            return

        for i, filepath in enumerate(self.pdf_files):
            card = FileCard(
                self.list_container,
                filename=os.path.basename(filepath),
                on_remove=lambda fp=filepath: self._remove_file(fp),
            )
            card.pack(fill="x", pady=(0, 4))

        count = len(self.pdf_files)
        self.count_label.configure(text=f"{count} file{'s' if count != 1 else ''}")
        self.extract_btn.configure(state="normal")

    def _add_files(self):
        files = filedialog.askopenfilenames(
            title="Select Invoice PDFs",
            filetypes=[("PDF Files", "*.pdf")],
        )
        for f in files:
            if f not in self.pdf_files:
                self.pdf_files.append(f)

        self._refresh_file_list()
        if self.pdf_files:
            self.status_label.configure(text=f"{len(self.pdf_files)} invoice(s) ready to process.")

    def _remove_file(self, filepath):
        if filepath in self.pdf_files:
            self.pdf_files.remove(filepath)
        self._refresh_file_list()

    def _clear_files(self):
        self.pdf_files.clear()
        self._refresh_file_list()

    def _extract(self):
        if not self.pdf_files:
            return

        self.status_label.configure(text="Processing invoices...", text_color=ACCENT)
        self.extract_btn.configure(state="disabled")
        self.update_idletasks()

        try:
            invoice_data = []
            company_names = set()
            for pdf_path in self.pdf_files:
                company, inv_num, refs = extract_references_from_pdf(pdf_path)
                company_names.add(company)
                invoice_data.append((inv_num, refs))

            invoice_data.sort(key=lambda x: x[0])

            company_name = max(company_names, key=lambda c: sum(1 for cn in company_names if cn == c))

            output_path = filedialog.asksaveasfilename(
                title="Save Spreadsheet As",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile="invoice_references.xlsx",
            )

            if not output_path:
                self.status_label.configure(text="Save cancelled.", text_color=TEXT_SECONDARY)
                self.extract_btn.configure(state="normal")
                return

            create_spreadsheet(company_name, invoice_data, output_path)

            total_refs = sum(len(refs) for _, refs in invoice_data)
            self.status_label.configure(
                text=f"Done! {total_refs} references from {len(invoice_data)} invoices.",
                text_color=SUCCESS,
            )
            self.extract_btn.configure(state="normal")
            messagebox.showinfo(
                "Success",
                f"Spreadsheet saved to:\n{output_path}\n\n"
                f"{len(invoice_data)} invoices processed\n"
                f"{total_refs} total references extracted",
            )
        except Exception as e:
            self.status_label.configure(text="Error occurred.", text_color=DANGER)
            self.extract_btn.configure(state="normal")
            messagebox.showerror("Error", f"Failed to process invoices:\n\n{str(e)}")


def main():
    app = InvoiceExtractorApp()
    app.mainloop()


if __name__ == "__main__":
    main()
